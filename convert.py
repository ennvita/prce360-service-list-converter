#!/usr/bin/env python3
"""Convert a telecom service list (Excel or Word) into the GridTemplate format."""

import sys
import re
from pathlib import Path
import openpyxl


# Tokens that indicate a company/organization name rather than a person name
_CORPORATE_TOKENS = frozenset({
    'llc', 'inc', 'corp', 'ltd', 'lp', 'llp', 'co', 'dba',
    'incorporated', 'corporation', 'limited', 'company',
    'communications', 'comms', 'telecom', 'telecommunications',
    'services', 'technologies', 'technology', 'solutions',
    'systems', 'network', 'networks', 'wireless', 'internet',
    'associates', 'enterprises', 'group', 'partners', 'partnership',
    'division',
})

# Email local parts that are clearly not person names
_GENERIC_LOCALS = frozenset({
    'compliance', 'regulatory', 'regulatorytax', 'accounting', 'tax', 'taxes',
    'info', 'payments', 'billing', 'admin', 'contact', 'support', 'sales',
    'accounts', 'mail', 'regdbg', 'starlinkregulatory', 'wirelinesalesandusetax',
    'indirect', 'bigstore', 'airvoice',
})

# Substrings in an email local that signal it's a company/dept address, not a person
_BUSINESS_SUBSTRINGS = frozenset({
    'technology', 'technologies', 'communications', 'wireless', 'telecom',
    'networks', 'solutions', 'services', 'systems', 'media', 'global',
    'financial', 'regulatory', 'compliance', 'billing', 'payments', 'accounts',
    'accounting', 'admin', 'contact', 'support', 'sales',
})

CHUNK_SIZE = 500


def _is_company_name(name: str) -> bool:
    if any(c.isdigit() for c in name):
        return True
    if '&' in name or ',' in name:
        return True
    tokens = [t.rstrip('.') for t in re.split(r'[\s./()]+', name.lower()) if t]
    return any(t in _CORPORATE_TOKENS for t in tokens)


def _parse_name_field(name: str) -> tuple[str, str]:
    name = name.strip()
    if not name:
        return '', ''
    parts = name.split(None, 1)
    first = parts[0].rstrip('.,')
    last = parts[1].rstrip('.,') if len(parts) > 1 else ''
    return first, last


def _try_person_from_local(local: str) -> tuple[str, str] | None:
    if not local:
        return None
    lower = local.lower()
    if lower in _GENERIC_LOCALS:
        return None
    if any(c.isdigit() for c in lower):
        return None
    if len(lower) > 18:
        return None
    if any(sub in lower for sub in _BUSINESS_SUBSTRINGS):
        return None

    if '.' in lower or (lower.isalpha() is False and '-' in lower):
        parts = [p for p in re.split(r'[.\-]', lower) if p and p.isalpha()]
        if len(parts) >= 2:
            return parts[0].capitalize(), parts[-1].capitalize()

    if '_' in lower:
        parts = [p for p in lower.split('_') if p and p.isalpha()]
        if len(parts) >= 2:
            return parts[0].capitalize(), parts[-1].capitalize()

    if not lower.isalpha():
        return None

    m = re.match(r'^([a-z])([a-z]{6,})$', lower)
    if m:
        return m.group(1).upper(), m.group(2).capitalize()

    if 2 <= len(lower) <= 15:
        return lower.capitalize(), ''

    return None


def _get_names(name: str | None, email: str) -> tuple[str, str]:
    local = ''
    domain_label = ''
    if email and '@' in email:
        local, remainder = email.split('@', 1)
        domain_label = remainder.split('.')[0]

    if name and not _is_company_name(name):
        first, last = _parse_name_field(name)
    else:
        person = _try_person_from_local(local) if local else None
        if person:
            first, last = person
        elif name:
            first, last = _parse_name_field(name)
        else:
            first, last = local.capitalize() or '', domain_label.capitalize() or ''

    if not last and domain_label:
        last = domain_label.capitalize()

    return first, last


def _clean_email(raw: object) -> str:
    addresses = [a.strip() for a in str(raw).split(';')]
    return next((a for a in addresses if a), '')


def _chunk_output_path(output_path: str, part: int, total_parts: int) -> str:
    if total_parts == 1:
        return output_path
    p = Path(output_path)
    return str(p.with_stem(f"{p.stem}_{part}"))


def _write_chunks(rows: list[tuple[str, str, str]], template_path: str, output_path: str) -> None:
    chunks = [rows[i:i + CHUNK_SIZE] for i in range(0, max(len(rows), 1), CHUNK_SIZE)]
    total_parts = len(chunks)

    for part, chunk in enumerate(chunks, start=1):
        out = _chunk_output_path(output_path, part, total_parts)
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        col = {h: idx + 1 for idx, h in enumerate(header) if h}

        for row_num, (first, last, email) in enumerate(chunk, start=2):
            ws.cell(row=row_num, column=col['First Name'], value=first)
            ws.cell(row=row_num, column=col['Last Name'], value=last)
            ws.cell(row=row_num, column=col['Type'], value='Individual')
            ws.cell(row=row_num, column=col['Email'], value=email)

        wb.save(out)
        print(f"Part {part}/{total_parts}: {len(chunk)} rows → '{out}'")


def convert(source_path: str, template_path: str, output_path: str) -> None:
    """Convert an Excel service list (.xlsx) to GridTemplate format."""
    source_wb = openpyxl.load_workbook(source_path)
    source_ws = source_wb.active

    rows: list[tuple[str, str, str]] = []
    skipped = 0

    for source_row in source_ws.iter_rows(min_row=2, values_only=True):
        raw_name, raw_email = source_row[0], source_row[1]
        email = _clean_email(raw_email) if raw_email else ''
        name = str(raw_name).strip() if raw_name else None

        if not name and not email:
            skipped += 1
            continue

        first, last = _get_names(name, email)
        rows.append((first, last, email))

    _write_chunks(rows, template_path, output_path)
    print(f"Done. {len(rows)} rows written, {skipped} blank rows skipped.")


def convert_docx(source_path: str, template_path: str, output_path: str) -> None:
    """Convert a Word service list (.docx) to GridTemplate format."""
    from docx import Document

    doc = Document(source_path)

    # Find the first table with Name and Email headers
    table = None
    for t in doc.tables:
        if not t.rows:
            continue
        headers = [c.text.strip().lower() for c in t.rows[0].cells]
        if any('name' in h for h in headers) and any('email' in h for h in headers):
            table = t
            break

    if table is None:
        print('Error: no Name/Email table found in the document.')
        sys.exit(1)

    rows: list[tuple[str, str, str]] = []
    skipped = 0
    _blank_prefix = re.compile(r'^[_\s]+')

    for row in table.rows[1:]:
        cells = [c.text.strip() for c in row.cells]
        raw_name = cells[0] if len(cells) > 0 else ''
        raw_email = cells[1] if len(cells) > 1 else ''

        email = _clean_email(raw_email) if raw_email else ''
        cleaned = _blank_prefix.sub('', raw_name).strip() if raw_name else ''
        name = cleaned or None

        if not name and not email:
            skipped += 1
            continue

        first, last = _get_names(name, email)
        rows.append((first, last, email))

    _write_chunks(rows, template_path, output_path)
    print(f"Done. {len(rows)} rows written, {skipped} blank rows skipped.")


if __name__ == '__main__':
    args = sys.argv[1:]
    if len(args) < 3:
        print("Usage: convert.py <source.xlsx|.docx> <template.xlsx> <output.xlsx>")
        sys.exit(1)

    source, template, output = args[0], args[1], args[2]

    for path in (source, template):
        if not Path(path).exists():
            print(f'Error: file not found: {path}')
            sys.exit(1)

    ext = Path(source).suffix.lower()
    if ext == '.xlsx':
        convert(source, template, output)
    elif ext == '.docx':
        convert_docx(source, template, output)
    else:
        print(f"Unsupported file type '{ext}'. Use .xlsx or .docx.")
        sys.exit(1)
